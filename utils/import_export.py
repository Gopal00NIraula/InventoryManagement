"""
Import/Export utilities for CSV and Excel files
Handles importing and exporting inventory items, suppliers, and customers
"""

import csv
import os
from datetime import datetime
from typing import List, Dict, Any, Tuple


def export_to_csv(data: List[Dict[str, Any]], filepath: str, headers: List[str]) -> Dict[str, Any]:
    """
    Export data to CSV file
    
    Args:
        data: List of dictionaries containing the data to export
        filepath: Path where the CSV file will be saved
        headers: List of column headers
    
    Returns:
        Dict with success status and message
    """
    try:
        with open(filepath, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=headers)
            writer.writeheader()
            
            for row in data:
                # Only write fields that are in headers
                filtered_row = {k: v for k, v in row.items() if k in headers}
                writer.writerow(filtered_row)
        
        return {
            "success": True,
            "message": f"Successfully exported {len(data)} records to {os.path.basename(filepath)}",
            "filepath": filepath
        }
    except Exception as e:
        return {
            "success": False,
            "message": f"Export failed: {str(e)}"
        }


def import_from_csv(filepath: str, required_headers: List[str]) -> Dict[str, Any]:
    """
    Import data from CSV file
    
    Args:
        filepath: Path to the CSV file to import
        required_headers: List of required column headers
    
    Returns:
        Dict with success status, data, and message
    """
    try:
        if not os.path.exists(filepath):
            return {
                "success": False,
                "message": "File not found"
            }
        
        data = []
        with open(filepath, 'r', encoding='utf-8') as csvfile:
            reader = csv.DictReader(csvfile)
            
            # Validate headers
            if not reader.fieldnames:
                return {
                    "success": False,
                    "message": "CSV file has no headers"
                }
            
            missing_headers = set(required_headers) - set(reader.fieldnames)
            if missing_headers:
                return {
                    "success": False,
                    "message": f"Missing required columns: {', '.join(missing_headers)}"
                }
            
            # Read data
            for row_num, row in enumerate(reader, start=2):
                # Validate required fields are not empty
                empty_fields = [h for h in required_headers if not row.get(h, '').strip()]
                if empty_fields:
                    return {
                        "success": False,
                        "message": f"Row {row_num}: Empty required fields: {', '.join(empty_fields)}"
                    }
                data.append(row)
        
        return {
            "success": True,
            "data": data,
            "message": f"Successfully loaded {len(data)} records from {os.path.basename(filepath)}"
        }
    except Exception as e:
        return {
            "success": False,
            "message": f"Import failed: {str(e)}"
        }


def export_inventory_to_csv(items: List[Dict[str, Any]], filepath: str) -> Dict[str, Any]:
    """Export inventory items to CSV"""
    headers = ['sku', 'name', 'quantity', 'price', 'min_stock_level', 'reorder_point', 'barcode']
    return export_to_csv(items, filepath, headers)


def import_inventory_from_csv(filepath: str) -> Dict[str, Any]:
    """Import inventory items from CSV"""
    required_headers = ['sku', 'name', 'quantity', 'price']
    result = import_from_csv(filepath, required_headers)
    
    if not result.get("success"):
        return result
    
    # Validate and convert data types
    validated_data = []
    for row_num, row in enumerate(result.get("data", []), start=2):
        try:
            # Convert numeric fields
            validated_row = row.copy()
            validated_row['quantity'] = int(row['quantity'])
            validated_row['price'] = float(row['price'])
            
            if row.get('min_stock_level'):
                validated_row['min_stock_level'] = int(row['min_stock_level'])
            else:
                validated_row['min_stock_level'] = 10
            
            if row.get('reorder_point'):
                validated_row['reorder_point'] = int(row['reorder_point'])
            else:
                validated_row['reorder_point'] = 20
            
            validated_data.append(validated_row)
        except ValueError as e:
            return {
                "success": False,
                "message": f"Row {row_num}: Invalid data format - {str(e)}"
            }
    
    result['data'] = validated_data
    return result


def export_suppliers_to_csv(suppliers: List[Dict[str, Any]], filepath: str) -> Dict[str, Any]:
    """Export suppliers to CSV"""
    headers = ['name', 'contact_person', 'email', 'phone', 'address']
    return export_to_csv(suppliers, filepath, headers)


def import_suppliers_from_csv(filepath: str) -> Dict[str, Any]:
    """Import suppliers from CSV"""
    required_headers = ['name', 'email', 'phone']
    return import_from_csv(filepath, required_headers)


def export_customers_to_csv(customers: List[Dict[str, Any]], filepath: str) -> Dict[str, Any]:
    """Export customers to CSV"""
    headers = ['name', 'email', 'phone', 'address']
    return export_to_csv(customers, filepath, headers)


def import_customers_from_csv(filepath: str) -> Dict[str, Any]:
    """Import customers from CSV"""
    required_headers = ['name', 'email', 'phone']
    return import_from_csv(filepath, required_headers)


try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    
    EXCEL_AVAILABLE = True
    
    def export_to_excel(data: List[Dict[str, Any]], filepath: str, headers: List[str], sheet_name: str = "Sheet1") -> Dict[str, Any]:
        """
        Export data to Excel file
        
        Args:
            data: List of dictionaries containing the data to export
            filepath: Path where the Excel file will be saved
            headers: List of column headers
            sheet_name: Name of the Excel sheet
        
        Returns:
            Dict with success status and message
        """
        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = sheet_name
            
            # Header styling
            header_fill = PatternFill(start_color="7c3aed", end_color="7c3aed", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            # Write headers
            for col_num, header in enumerate(headers, start=1):
                cell = sheet.cell(row=1, column=col_num, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Write data
            for row_num, row_data in enumerate(data, start=2):
                for col_num, header in enumerate(headers, start=1):
                    value = row_data.get(header, '')
                    sheet.cell(row=row_num, column=col_num, value=value)
            
            # Auto-adjust column widths
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                sheet.column_dimensions[column_letter].width = adjusted_width
            
            workbook.save(filepath)
            
            return {
                "success": True,
                "message": f"Successfully exported {len(data)} records to {os.path.basename(filepath)}",
                "filepath": filepath
            }
        except Exception as e:
            return {
                "success": False,
                "message": f"Export failed: {str(e)}"
            }
    
    
    def import_from_excel(filepath: str, required_headers: List[str], sheet_name: str = None) -> Dict[str, Any]:
        """
        Import data from Excel file
        
        Args:
            filepath: Path to the Excel file to import
            required_headers: List of required column headers
            sheet_name: Name of the sheet to import (None for active sheet)
        
        Returns:
            Dict with success status, data, and message
        """
        try:
            if not os.path.exists(filepath):
                return {
                    "success": False,
                    "message": "File not found"
                }
            
            workbook = openpyxl.load_workbook(filepath)
            sheet = workbook[sheet_name] if sheet_name else workbook.active
            
            # Get headers from first row
            headers = []
            for cell in sheet[1]:
                if cell.value:
                    headers.append(str(cell.value).strip())
            
            if not headers:
                return {
                    "success": False,
                    "message": "Excel file has no headers"
                }
            
            # Validate required headers
            missing_headers = set(required_headers) - set(headers)
            if missing_headers:
                return {
                    "success": False,
                    "message": f"Missing required columns: {', '.join(missing_headers)}"
                }
            
            # Read data
            data = []
            for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if not any(row):  # Skip empty rows
                    continue
                
                row_dict = {headers[i]: (row[i] if i < len(row) else '') for i in range(len(headers))}
                
                # Validate required fields
                empty_fields = [h for h in required_headers if not str(row_dict.get(h, '')).strip()]
                if empty_fields:
                    return {
                        "success": False,
                        "message": f"Row {row_num}: Empty required fields: {', '.join(empty_fields)}"
                    }
                
                data.append(row_dict)
            
            return {
                "success": True,
                "data": data,
                "message": f"Successfully loaded {len(data)} records from {os.path.basename(filepath)}"
            }
        except Exception as e:
            return {
                "success": False,
                "message": f"Import failed: {str(e)}"
            }
    
    
    def export_inventory_to_excel(items: List[Dict[str, Any]], filepath: str) -> Dict[str, Any]:
        """Export inventory items to Excel"""
        headers = ['sku', 'name', 'quantity', 'price', 'min_stock_level', 'reorder_point', 'barcode']
        return export_to_excel(items, filepath, headers, "Inventory")
    
    
    def import_inventory_from_excel(filepath: str) -> Dict[str, Any]:
        """Import inventory items from Excel"""
        required_headers = ['sku', 'name', 'quantity', 'price']
        result = import_from_excel(filepath, required_headers)
        
        if not result.get("success"):
            return result
        
        # Validate and convert data types (same as CSV)
        validated_data = []
        for row_num, row in enumerate(result.get("data", []), start=2):
            try:
                validated_row = row.copy()
                validated_row['quantity'] = int(float(str(row['quantity'])))
                validated_row['price'] = float(row['price'])
                
                if row.get('min_stock_level') and str(row.get('min_stock_level')).strip():
                    validated_row['min_stock_level'] = int(float(str(row['min_stock_level'])))
                else:
                    validated_row['min_stock_level'] = 10
                
                if row.get('reorder_point') and str(row.get('reorder_point')).strip():
                    validated_row['reorder_point'] = int(float(str(row['reorder_point'])))
                else:
                    validated_row['reorder_point'] = 20
                
                validated_data.append(validated_row)
            except ValueError as e:
                return {
                    "success": False,
                    "message": f"Row {row_num}: Invalid data format - {str(e)}"
                }
        
        result['data'] = validated_data
        return result
    
    
    def export_suppliers_to_excel(suppliers: List[Dict[str, Any]], filepath: str) -> Dict[str, Any]:
        """Export suppliers to Excel"""
        headers = ['name', 'contact_person', 'email', 'phone', 'address']
        return export_to_excel(suppliers, filepath, headers, "Suppliers")
    
    
    def import_suppliers_from_excel(filepath: str) -> Dict[str, Any]:
        """Import suppliers from Excel"""
        required_headers = ['name', 'email', 'phone']
        return import_from_excel(filepath, required_headers)
    
    
    def export_customers_to_excel(customers: List[Dict[str, Any]], filepath: str) -> Dict[str, Any]:
        """Export customers to Excel"""
        headers = ['name', 'email', 'phone', 'address']
        return export_to_excel(customers, filepath, headers, "Customers")
    
    
    def import_customers_from_excel(filepath: str) -> Dict[str, Any]:
        """Import customers from Excel"""
        required_headers = ['name', 'email', 'phone']
        return import_from_excel(filepath, required_headers)

except ImportError:
    EXCEL_AVAILABLE = False
    
    def excel_not_available(*args, **kwargs):
        return {
            "success": False,
            "message": "Excel support not available. Install openpyxl: pip install openpyxl"
        }
    
    export_to_excel = excel_not_available
    import_from_excel = excel_not_available
    export_inventory_to_excel = excel_not_available
    import_inventory_from_excel = excel_not_available
    export_suppliers_to_excel = excel_not_available
    import_suppliers_from_excel = excel_not_available
    export_customers_to_excel = excel_not_available
    import_customers_from_excel = excel_not_available
